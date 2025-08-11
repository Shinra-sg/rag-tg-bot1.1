#!/usr/bin/env python3
"""
Парсер для Excel файлов (.xlsx, .xls)
"""

import sys
import json
import pandas as pd
from openpyxl import load_workbook
import os

def parse_excel(file_path):
    """Парсит Excel файл и возвращает текст и метаданные"""
    try:
        # Загружаем файл
        workbook = load_workbook(file_path, data_only=True)
        
        content_parts = []
        metadata = {
            "title": os.path.basename(file_path),
            "sheets": len(workbook.sheetnames),
            "created": None,
            "modified": None
        }
        
        # Обрабатываем каждый лист
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content_parts.append(f"=== Лист: {sheet_name} ===")
            
            # Получаем данные из листа
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    content_parts.append(row_text)
            
            content_parts.append("")  # Пустая строка между листами
        
        content = "\n".join(content_parts)
        
        # Если контент пустой, пробуем pandas
        if not content.strip():
            try:
                df = pd.read_excel(file_path, sheet_name=None)
                content_parts = []
                
                for sheet_name, sheet_df in df.items():
                    content_parts.append(f"=== Лист: {sheet_name} ===")
                    content_parts.append(sheet_df.to_string(index=False))
                    content_parts.append("")
                
                content = "\n".join(content_parts)
            except Exception as e:
                content = f"Ошибка чтения Excel файла: {str(e)}"
        
        return {
            "content": content,
            "metadata": metadata
        }
        
    except Exception as e:
        return {
            "content": f"Ошибка парсинга Excel файла: {str(e)}",
            "metadata": {"title": os.path.basename(file_path)}
        }

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(json.dumps({"error": "Необходимо указать путь к файлу"}))
        sys.exit(1)
    
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(json.dumps({"error": f"Файл не найден: {file_path}"}))
        sys.exit(1)
    
    result = parse_excel(file_path)
    print(json.dumps(result, ensure_ascii=False, default=str)) 